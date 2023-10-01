import pandas as pd
from openpyxl import load_workbook, Workbook



def extract_info_from_excel(file_path):
    # Load the Excel file
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Convert worksheet to dataframe for easier parsing
    df = pd.DataFrame(ws.values)

    names, professions, companies, locations, experiences, hyperlinks = [], [], [], [], [], []

    i = 0
    while i < len(df):
        line = df.iloc[i, 0]

        # Check if the current line is a name based on the assumption that the next line will contain the connection degree
        if i + 1 < len(df) and any(deg in str(df.iloc[i + 1, 0]) for deg in ["1st", "2nd", "3rd"]):
            name = line
            hyperlink = ws.cell(row=i + 1, column=1).hyperlink.target if ws.cell(row=i + 1,
                                                                                 column=1).hyperlink else None

            # Get profession and company
            i += 2  # Skip the connection info and get to the profession
            profession_company = df.iloc[i, 0].split("  ") if df.iloc[i, 0] else [None, None]
            profession, company = (profession_company + [None, None])[:2]

            # Get location
            i += 1
            location = df.iloc[i, 0] if df.iloc[i, 0] else None

            # Find experience
            exp_line = ""
            j = i + 1

            while j < len(df) and not str(df.iloc[j, 0]).startswith("Experience:"):
                j += 1
            if j < len(df) and str(df.iloc[j, 0]).startswith("Experience:"):
                exp_line = df.iloc[j, 0].replace("Experience:", "").strip()
                i = j

            # Append the data
            names.append(name)
            professions.append(profession)
            companies.append(company)
            locations.append(location)
            experiences.append(exp_line)
            hyperlinks.append(hyperlink)

        i += 1

    # Convert to DataFrame for processing
    data_df = pd.DataFrame({
        'Name': names,
        'Profession': professions,
        'Companies': companies,
        'Locations': locations,
        'Experience': experiences
    })

    # Create a new Excel workbook and sheet
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.append(["Name", "Profession", "Companies", "Locations", "Experience"])

    for i, row in data_df.iterrows():
        ws_out.append(row.tolist())

        # If there's a hyperlink in the original data, assign it to the name in the output file
        if hyperlinks[i]:
            ws_out.cell(row=ws_out.max_row, column=1).hyperlink = hyperlinks[i]

    wb_out.save('/Users/jakub/Downloads/Output_with_links.xlsx')
    return data_df


if __name__ == "__main__":
    extract_info_from_excel('/Users/jakub/Downloads/list.xlsx')
