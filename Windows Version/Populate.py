import pandas as pd
from openpyxl import load_workbook
import re


def map_to_company_url(swiss_connection):
    company_urls_mapping = {
        "roche": "https://www.roche.com",
        "novartis": "https://www.novartis.com",
    }
    urls = [company_urls_mapping[company.lower()] for company in swiss_connection.split(", ") if
            company.lower() in company_urls_mapping]
    return ", ".join(urls) if urls else "None"


def get_swiss_connection(exp):
    pattern = r'\)\s+([^ ]+)\s+'
    matches = re.findall(pattern, exp)
    unique_companies = set(matches)
    return ", ".join(unique_companies) if unique_companies else "None"


def calculate_experience_in_months(experience):
    pattern = r'(\d+)\s*yr|(\d+)\s*mo'
    matches = re.findall(pattern, experience)
    total_months = 0
    for year, month in matches:
        if year:
            total_months += int(year) * 12
        if month:
            total_months += int(month)
    return total_months


def append_to_spreadsheet(source_file, existing_file):
    # Load the data from the source file
    wb_links = load_workbook(source_file, data_only=True)
    ws_links = wb_links.active
    df_input = pd.read_excel(source_file)
    df_input['Original Experience'] = df_input['Experience']
    df_input['Experience'] = df_input['Experience'].apply(calculate_experience_in_months)

    swiss_companies_pattern = re.compile(r'\(\s*\d+\s+(yrs?|mos?)\s*\d*\s*(yrs?|mos?)?\s*\)\s*([^ ]+)\s*')

    df_input['Swiss Connection'] = df_input['Original Experience'].apply(get_swiss_connection)

    # Extract LinkedIn URLs from the hyperlinks in the 'Name' column
    linkedin_urls = []
    for i, row in enumerate(ws_links.iter_rows(min_row=2, max_row=ws_links.max_row, min_col=1, max_col=1)):
        hyperlink = row[0].hyperlink
        if hyperlink is not None:
            linkedin_urls.append(hyperlink.target)
        else:
            linkedin_urls.append(None)

    # Load the existing Excel workbook and worksheet
    wb_existing = load_workbook(existing_file)
    ws = wb_existing.active

    # Convert the existing Excel data to a DataFrame for easier comparison
    data = list(ws.iter_rows(values_only=True))
    columns = data[0]
    df_existing = pd.DataFrame(data[1:], columns=columns)

    # Match on Name and Full Name columns to get unique rows
    merged_df = df_input.merge(df_existing, left_on="Name", right_on="Full Name", how='left', indicator=True)
    unique_rows = merged_df[merged_df['_merge'] == 'left_only'].copy()

    # Map values to the existing spreadsheet format

    unique_rows['Location'] = unique_rows['Locations']
    unique_rows['Source (url)'] = "Linkedin Sales Navigator Tool"
    unique_rows['Full Name'] = unique_rows['Name']
    unique_rows['Title'] = unique_rows['Profession']
    unique_rows['Company'] = unique_rows['Companies']
    linkedin_urls_filtered = [linkedin_urls[i] for i in unique_rows.index]
    unique_rows['LinkedIn (url)'] = linkedin_urls_filtered
    unique_rows.loc[:, 'Experience'] = df_input.loc[unique_rows.index, 'Experience']

    # Drop the merge column and any other unneeded columns
    unique_rows = unique_rows.drop(columns=['_merge', 'Name', 'Profession', 'Locations'])

    columns_order = ['Location', 'Swiss Connection', 'Source (url)', 'Full Name', 'Title', 'Company',
                     'Experience', 'LinkedIn (url)']
    unique_rows.rename(columns={'Swiss Connection_x': 'Swiss Connection'}, inplace=True)

    unique_rows = unique_rows[columns_order]

    # Find the starting row to append data in existing_file
    start_row = ws.max_row + 1

    # Initialize a set to store existing URLs
    existing_urls = set(df_existing['LinkedIn (url)'])

    # Append unique rows to the existing spreadsheet
    for index, row in unique_rows.iterrows():
        existing_url = linkedin_urls[index]
        if existing_url is not None and existing_url not in existing_urls:
            # Check if any data in the row already exists in the existing sheet
            duplicate_row = False
            for _, existing_row in df_existing.iterrows():
                if all(existing_row[col] == row[col] for col in columns_order):
                    duplicate_row = True
                    break

            if duplicate_row:
                continue

            existing_urls.add(existing_url)

            for c_idx, value in enumerate(row, 1):
                ws.cell(row=start_row, column=c_idx, value=value)
                if c_idx == 7 and isinstance(value, str) and value.startswith("http"):
                    ws.cell(row=start_row, column=c_idx).hyperlink = value
                if c_idx == 8 and isinstance(value, str) and value.startswith("http"):
                    original_link_formula = ws_links.cell(row=index + 2, column=1).hyperlink.target
                    ws.cell(row=start_row, column=c_idx).hyperlink = original_link_formula
            start_row += 1

    # Save the modified workbook
    wb_existing.save(existing_file)


if __name__ == "__main__":
    source_file = '/Users/jakub/Downloads/Output_with_links.xlsx'
    existing_file = '/Users/jakub/Downloads/Existing_spreadsheet.xlsx'
    append_to_spreadsheet(source_file, existing_file)
